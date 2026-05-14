from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io, os

def generer_pdf(mission, collectes):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    BLEU  = colors.HexColor("#1a3a6b")
    VERT  = colors.HexColor("#1a6b3c")
    GRIS  = colors.HexColor("#f7fafc")
    GRIS2 = colors.HexColor("#e2e8f0")

    elements = []

    # TITRE
    elements.append(Paragraph("DataBroker 229", ParagraphStyle(
        "t", fontSize=20, textColor=BLEU,
        fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=4)))
    elements.append(Paragraph("Rapport de collecte de données terrain", ParagraphStyle(
        "s", fontSize=11, textColor=VERT,
        fontName="Helvetica", alignment=TA_CENTER, spaceAfter=2)))
    elements.append(Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
        ParagraphStyle("d", fontSize=9, textColor=colors.grey,
                       alignment=TA_CENTER, spaceAfter=16)))

    elements.append(Table([[""]], colWidths=[17*cm],
        style=TableStyle([("LINEBELOW",(0,0),(-1,-1),2,BLEU)])))
    elements.append(Spacer(1, 0.4*cm))

    # INFOS MISSION
    elements.append(Paragraph("Détails de la mission", ParagraphStyle(
        "sec", fontSize=12, textColor=BLEU,
        fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=8)))

    info_data = [
        ["Mission",   mission.titre],
        ["Marché",    mission.marche_cible or "—"],
        ["Produit",   mission.produit_cible or "—"],
        ["Statut",    mission.statut.upper()],
        ["Date",      mission.date_creation.strftime("%d/%m/%Y")],
        ["Collectes", str(len(collectes))],
    ]
    t = Table(info_data, colWidths=[5*cm, 12*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",     (0,0),(0,-1), GRIS),
        ("TEXTCOLOR",      (0,0),(0,-1), BLEU),
        ("FONTNAME",       (0,0),(0,-1), "Helvetica-Bold"),
        ("FONTSIZE",       (0,0),(-1,-1), 9),
        ("GRID",           (0,0),(-1,-1), 0.5, GRIS2),
        ("ROWBACKGROUNDS", (0,0),(-1,-1), [colors.white, GRIS]),
        ("PADDING",        (0,0),(-1,-1), 8),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.5*cm))

    # STATISTIQUES
    elements.append(Paragraph("Statistiques", ParagraphStyle(
        "sec2", fontSize=12, textColor=BLEU,
        fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=8)))

    prix_vals = [c.prix_observe for c in collectes if c.prix_observe]
    nb_dispo  = sum(1 for c in collectes if c.disponibilite)

    prix_moy   = f"{sum(prix_vals)/len(prix_vals):,.0f} FCFA" if prix_vals else "—"
    prix_min   = f"{min(prix_vals):,.0f} FCFA" if prix_vals else "—"
    prix_max   = f"{max(prix_vals):,.0f} FCFA" if prix_vals else "—"
    taux_dispo = f"{(nb_dispo/len(collectes)*100):.0f}%" if collectes else "—"

    stats = Table(
        [["Prix moyen","Prix minimum","Prix maximum","Disponibilité"],
         [prix_moy, prix_min, prix_max, taux_dispo]],
        colWidths=[4.25*cm]*4)
    stats.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,0), BLEU),
        ("TEXTCOLOR",  (0,0),(-1,0), colors.white),
        ("FONTNAME",   (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTNAME",   (0,1),(-1,1), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0),(-1,-1), 9),
        ("ALIGN",      (0,0),(-1,-1), "CENTER"),
        ("VALIGN",     (0,0),(-1,-1), "MIDDLE"),
        ("BACKGROUND", (0,1),(-1,1), colors.HexColor("#eaf4ee")),
        ("TEXTCOLOR",  (0,1),(-1,1), VERT),
        ("GRID",       (0,0),(-1,-1), 0.5, GRIS2),
        ("ROWHEIGHT",  (0,0),(-1,-1), 28),
    ]))
    elements.append(stats)
    elements.append(Spacer(1, 0.5*cm))

    # DONNÉES INDIVIDUELLES
    elements.append(Paragraph("Données individuelles", ParagraphStyle(
        "sec3", fontSize=12, textColor=BLEU,
        fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=8)))

    rows = [["#","Date","Prix (FCFA)","Disponible","Commentaire","GPS"]]
    for i, c in enumerate(collectes, 1):
        gps = f"{c.latitude:.4f}, {c.longitude:.4f}" if c.latitude else "—"
        rows.append([
            str(i),
            c.date_soumission.strftime("%d/%m/%Y %H:%M"),
            f"{c.prix_observe:,.0f}" if c.prix_observe else "—",
            "Oui" if c.disponibilite else "Non",
            (c.commentaire[:50]+"...") if c.commentaire and len(c.commentaire)>50 else (c.commentaire or "—"),
            gps
        ])

    dt = Table(rows, colWidths=[1*cm,3*cm,3*cm,2.5*cm,5.5*cm,3*cm], repeatRows=1)
    style_dt = [
        ("BACKGROUND",     (0,0),(-1,0), BLEU),
        ("TEXTCOLOR",      (0,0),(-1,0), colors.white),
        ("FONTNAME",       (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",       (0,0),(-1,-1), 8),
        ("ALIGN",          (0,0),(-1,-1), "CENTER"),
        ("VALIGN",         (0,0),(-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1),(-1,-1), [colors.white, GRIS]),
        ("GRID",           (0,0),(-1,-1), 0.4, GRIS2),
        ("PADDING",        (0,0),(-1,-1), 6),
    ]
    for i in range(1, len(rows)):
        couleur = VERT if rows[i][3] == "Oui" else colors.red
        style_dt.append(("TEXTCOLOR", (3,i), (3,i), couleur))
    dt.setStyle(TableStyle(style_dt))
    elements.append(dt)
    elements.append(Spacer(1, 0.5*cm))

    # PHOTOS DE TERRAIN
    photos = [c for c in collectes if c.photo_url]
    if photos:
        elements.append(Paragraph("Photos de terrain", ParagraphStyle(
            "sec4", fontSize=12, textColor=BLEU,
            fontName="Helvetica-Bold", spaceBefore=12, spaceAfter=8)))
        for i, c in enumerate(photos, 1):
            chemin = c.photo_url.lstrip("/")
            if os.path.exists(chemin):
                try:
                    img = Image(chemin, width=8*cm, height=6*cm)
                    legende = Paragraph(
                        f"Photo {i} — {c.date_soumission.strftime('%d/%m/%Y %H:%M')} "
                        f"| Prix: {c.prix_observe or '—'} FCFA",
                        ParagraphStyle("leg", fontSize=8,
                                       textColor=colors.grey, alignment=TA_CENTER))
                    elements.append(img)
                    elements.append(legende)
                    elements.append(Spacer(1, 0.3*cm))
                except Exception:
                    pass

    # PIED DE PAGE
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Table([[""]], colWidths=[17*cm],
        style=TableStyle([("LINEABOVE",(0,0),(-1,-1),1,GRIS2)])))
    elements.append(Paragraph(
        "DataBroker 229 — Les yeux des entreprises sur les marchés béninois",
        ParagraphStyle("f", fontSize=8, textColor=colors.grey, alignment=TA_CENTER)))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generer_excel(mission, collectes):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport"

    BLEU  = "1A3A6B"
    VERT  = "1A6B3C"
    GRIS  = "F7FAFC"
    GRIS2 = "E2E8F0"
    BLANC = "FFFFFF"

    def style(cell, bold=False, bg=None, fg="000000", size=10, align="left"):
        cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(
            horizontal=align, vertical="center", wrap_text=True)

    thin   = Side(style="thin", color=GRIS2)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # TITRE
    ws.merge_cells("A1:G1")
    ws["A1"] = "DATABROKER 229 — Rapport de collecte terrain"
    style(ws["A1"], bold=True, bg=BLEU, fg=BLANC, size=14, align="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    style(ws["A2"], bg=GRIS, fg="718096", align="center")

    # INFOS MISSION
    ws.merge_cells("A4:G4")
    ws["A4"] = "DÉTAILS DE LA MISSION"
    style(ws["A4"], bold=True, bg=BLEU, fg=BLANC, align="center")

    infos = [
        ("Mission",   mission.titre),
        ("Marché",    mission.marche_cible or "—"),
        ("Produit",   mission.produit_cible or "—"),
        ("Statut",    mission.statut.upper()),
        ("Date",      mission.date_creation.strftime("%d/%m/%Y")),
        ("Collectes", str(len(collectes))),
    ]
    for i, (label, valeur) in enumerate(infos, start=5):
        ws[f"A{i}"] = label
        style(ws[f"A{i}"], bold=True, bg=GRIS, fg=BLEU)
        ws.merge_cells(f"B{i}:G{i}")
        ws[f"B{i}"] = valeur
        style(ws[f"B{i}"])
        ws.row_dimensions[i].height = 20

    # STATISTIQUES
    r = len(infos) + 6
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "STATISTIQUES"
    style(ws[f"A{r}"], bold=True, bg=VERT, fg=BLANC, align="center")

    prix_vals = [c.prix_observe for c in collectes if c.prix_observe]
    nb_dispo  = sum(1 for c in collectes if c.disponibilite)

    slabels = ["Prix moyen","Prix minimum","Prix maximum","Taux dispo","Nb collectes"]
    svals   = [
        f"{sum(prix_vals)/len(prix_vals):,.0f} FCFA" if prix_vals else "—",
        f"{min(prix_vals):,.0f} FCFA" if prix_vals else "—",
        f"{max(prix_vals):,.0f} FCFA" if prix_vals else "—",
        f"{(nb_dispo/len(collectes)*100):.0f}%" if collectes else "—",
        str(len(collectes))
    ]
    for ci, (sl, sv) in enumerate(zip(slabels, svals), start=1):
        cl = ws.cell(row=r+1, column=ci, value=sl)
        cv = ws.cell(row=r+2, column=ci, value=sv)
        style(cl, bold=True, bg=GRIS, fg=BLEU, align="center")
        style(cv, bold=True, fg=VERT, align="center")
        cv.font = Font(bold=True, color=VERT, size=12, name="Calibri")

    # DONNÉES
    rd = r + 4
    ws.merge_cells(f"A{rd}:G{rd}")
    ws[f"A{rd}"] = "DONNÉES INDIVIDUELLES"
    style(ws[f"A{rd}"], bold=True, bg=BLEU, fg=BLANC, align="center")

    headers = ["#","Date","Prix (FCFA)","Disponible","Commentaire","GPS Lat","GPS Lng"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=rd+1, column=ci, value=h)
        style(c, bold=True, bg="2D5F9E", fg=BLANC, align="center")
        c.border = border

    for i, col in enumerate(collectes, start=1):
        rw = rd + 1 + i
        bg = BLANC if i % 2 == 0 else GRIS
        vals = [
            i,
            col.date_soumission.strftime("%d/%m/%Y %H:%M"),
            col.prix_observe or "—",
            "Oui" if col.disponibilite else "Non",
            col.commentaire or "—",
            col.latitude or "—",
            col.longitude or "—",
        ]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=rw, column=ci, value=v)
            style(c, bg=bg, align="center")
            c.border = border
            if ci == 4:
                c.font = Font(bold=True, name="Calibri", size=10,
                              color=VERT if v == "Oui" else "C53030")
        ws.row_dimensions[rw].height = 18

    largeurs = [5, 18, 14, 12, 40, 15, 15]
    for ci, larg in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = larg

    ws.freeze_panes = f"A{rd+2}"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
